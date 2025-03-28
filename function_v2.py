import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
import streamlit as st
import plotly.graph_objects as go
import plotly.express as px

# =========================================================
# Function to generate index file in a certain format
# =========================================================
# Define the function to process the data
def process_index_data(file_path, portfolio_name):
    # Read the CSV file
    index_file = pd.read_csv(file_path, delimiter=';')
    
    # Select relevant columns and convert 'Date' to datetime format
    index_file2 = index_file[['Date', 'Indexvalue']]
    index_file2['Date'] = pd.to_datetime(index_file2['Date'])
    
    # Rename the columns based on the provided portfolio name
    index_file2.rename(columns={
        'Indexvalue': f'Indexvalue_{portfolio_name}'
    }, inplace=True)

    # Find the last date in the dataset
    last_date = index_file2['Date'].max()
    
    # Identify the last available month and year
    last_month = last_date.month
    last_year = last_date.year
    
    # Identify the next month (next_month will be the current month + 1)
    next_month = last_month + 1 if last_month < 12 else 1
    next_year = last_year if last_month < 12 else last_year + 1
    
    # Check if there is any data for the next month (this would indicate if the current month is incomplete)
    next_month_data = index_file2[(index_file2['Date'].dt.month == next_month) & (index_file2['Date'].dt.year == next_year)]
    
    # If data for the next month exists, we consider the current month as incomplete
    if len(next_month_data) == 0:  # No data for next month means the current month is incomplete
        incomplete_month = last_month
    else:
        incomplete_month = None
       
    # Remove the data for the incomplete month if necessary
    if incomplete_month:
        index_file2 = index_file2[~((index_file2['Date'].dt.month == incomplete_month) & (index_file2['Date'].dt.year == last_year))]
    return index_file2




# =========================================================
# Function to get previous Available dates from the datafile
# =========================================================
## This function is needed when we need a start index value for return calculation. The start date is usually the target period's start date- 1 day
def get_previous_available_date(start_date, df):
    # Ensure the 'Date' column is in datetime format
    # df['Date'] = pd.to_datetime(df['Date'])
    
    # Check if the start_date exists in the dataset
    while start_date not in df['Date'].values:
        # If the date doesn't exist, subtract 1 day and check again
        start_date -= pd.Timedelta(days=1)
        
        # If no dates are available before the given date, break
        if start_date < df['Date'].min():
            return None  # No previous date available
        
    return start_date


# =====================================================
# Function to get NEXT Available dates from the datafile
# ======================================================
## This function is needed when Today - Period days is not a valid date. then it should pick the next available date. 
def get_closest_date(start_date, df):
    # Filter out all dates that are on or after the start_date
    valid_dates = df[df['Date'] >= start_date]
    
    # If there are no dates on or after the start_date, return None
    if valid_dates.empty:
        return None
    
    # Check if the start_date itself is present
    if start_date in valid_dates['Date'].values:
        return start_date
    else:
        # If the start_date is not available, return the next available date
        return valid_dates.iloc[0]['Date']



# ========================================================
# Function to check if any year in the past was a leap year
# =========================================================
# Function to check if a year is a leap year
def is_leap_year(year):
    return (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0))


# ============================================================================================
# Function to get period dates for each of the return type  (YTD, 1Month, 1yr, etc) calculation
# ============================================================================================
# Function to identify the start and end dates for all periods and store them in a DataFrame
def get_period_dates(df):
    # Ensure the 'Date' column is in datetime format
    df['Date'] = pd.to_datetime(df['Date'])
    
    # Get the latest date in the dataset (end date)
    end_date = df['Date'].max()
    
    # Initialize a list to store period information
    period_dates = []

    # 1-Month Start and End Date
    start_date_1m = end_date.replace(day=1)  # Start date should be the first day of the current month
    actual_start_date_1m = get_closest_date(start_date_1m, df)
    last_value_date_1m=actual_start_date_1m-pd.DateOffset(days=1)
    last_value_date_1m = get_previous_available_date(last_value_date_1m, df)
    period_dates.append({"Period": "1 Month", "Last Value Date": last_value_date_1m, "Start Date": actual_start_date_1m, "End Date": end_date})
    
    # YTD Start and End Date
    start_date_ytd = pd.to_datetime(f"{end_date.year}-01-01")
    actual_start_date_ytd = get_closest_date(start_date_ytd, df)
    last_value_date_ytd=actual_start_date_ytd-pd.DateOffset(days=1)
    last_value_date_ytd = get_previous_available_date(last_value_date_ytd, df)
    period_dates.append({"Period": "YTD", "Last Value Date": last_value_date_ytd, "Start Date": actual_start_date_ytd, "End Date": end_date})


    # 1-Year Start and End Date (considering leap year)
    start_date_1yr = end_date - pd.DateOffset(years=1) + pd.DateOffset(days=1)  # Adjust to next day after 1 year back
    if is_leap_year(end_date.year - 1) and end_date.month == 2 and end_date.day == 28:
        start_date_1yr = start_date_1yr.replace(month=3, day=1)  # If the previous year was a leap year, set to March 1
    actual_start_date_1yr = get_closest_date(start_date_1yr, df)
    last_value_date_1yr = actual_start_date_1yr - pd.DateOffset(days=1)
    last_value_date_1yr = get_previous_available_date(last_value_date_1yr, df)
    period_dates.append({"Period": "1 Year", "Last Value Date": last_value_date_1yr, "Start Date": actual_start_date_1yr, "End Date": end_date})

    # 3-Year Start and End Date (considering leap year)
    start_date_3yr = end_date - pd.DateOffset(years=3) + pd.DateOffset(days=1)  # Adjust to next day after 3 years back
    if is_leap_year(end_date.year - 3) and end_date.month == 2 and end_date.day == 28:
        start_date_3yr = start_date_3yr.replace(month=3, day=1)  # If the previous year was a leap year, set to March 1
    actual_start_date_3yr = get_closest_date(start_date_3yr, df)
    last_value_date_3yr = actual_start_date_3yr - pd.DateOffset(days=1)
    last_value_date_3yr = get_previous_available_date(last_value_date_3yr, df)
    period_dates.append({"Period": "3 Years", "Last Value Date": last_value_date_3yr, "Start Date": actual_start_date_3yr, "End Date": end_date})

    # 5-Year Start and End Date (considering leap year)
    start_date_5yr = end_date - pd.DateOffset(years=5) + pd.DateOffset(days=1)  # Adjust to next day after 5 years back
    if is_leap_year(end_date.year - 5) and end_date.month == 2 and end_date.day == 28:
        start_date_5yr = start_date_5yr.replace(month=3, day=1)  # If the previous year was a leap year, set to March 1
    actual_start_date_5yr = get_closest_date(start_date_5yr, df)
    last_value_date_5yr = actual_start_date_5yr - pd.DateOffset(days=1)
    last_value_date_5yr = get_previous_available_date(last_value_date_5yr, df)
    period_dates.append({"Period": "5 Years", "Last Value Date": last_value_date_5yr, "Start Date": actual_start_date_5yr, "End Date": end_date})

    # 10-Year Start and End Date (considering leap year)
    start_date_10yr = end_date - pd.DateOffset(years=10) + pd.DateOffset(days=1)  # Adjust to next day after 3 years back
    if is_leap_year(end_date.year - 3) and end_date.month == 2 and end_date.day == 28:
        start_date_10yr = start_date_10yr.replace(month=3, day=1)  # If the previous year was a leap year, set to March 1
    actual_start_date_10yr = get_closest_date(start_date_10yr, df)
    last_value_date_10yr = actual_start_date_10yr - pd.DateOffset(days=1)
    last_value_date_10yr = get_previous_available_date(last_value_date_10yr, df)
    period_dates.append({"Period": "10 Years", "Last Value Date": last_value_date_10yr, "Start Date": actual_start_date_10yr, "End Date": end_date})

    # Convert the list of dictionaries to a DataFrame
    period_df = pd.DataFrame(period_dates)
    
    return period_df

# ============================================
# Function to calculate annualised return
# ============================================
### To calculate annualised return formula. 
def annualised_return(start_value, end_value, start_date, end_date, df):
    # Filter the DataFrame to get only the relevant dates between start_date and end_date
    relevant_data = df[(df['Date'] >= start_date) & (df['Date'] <= end_date)]
    
    # Count the number of days in the dataset (actual trading days)
    period_days = len(relevant_data)
    
    # Calculate simple return
    simple_return = (end_value - start_value) / start_value
    
    # Apply the annualising formula
    annualised_return = (1 + simple_return) ** (260 / period_days) - 1
    return annualised_return

# ============================================
# Function to calculate volatility
# ============================================
## Function to calculate Volatility formula
import numpy as np

def calculate_volatility(start_date, end_date, df):
    # Dynamically find the Return column that starts with 'Return_'
    return_column_name = [col for col in df.columns if col.startswith('Return_')]
    
    if not return_column_name:
        raise ValueError("No Return column found in the dataframe")  # Raise an error if no return column exists
    
    # Only take the first return column found (assuming there's only one)
    return_column_name = return_column_name[0]
    
    # Filter the DataFrame to get only the relevant dates between start_date and end_date
    relevant_data = df[(df['Date'] >= start_date) & (df['Date'] <= end_date)]
    
    if len(relevant_data) < 2:
        return np.nan  # Not enough data to calculate volatility
    
    # Drop the rows with NaN in the return column
    relevant_data = relevant_data.dropna(subset=[return_column_name])
    
    # Calculate the sample variance of daily returns
    sample_variance = relevant_data[return_column_name].var()
    
    # Calculate the standard deviation (volatility) as the square root of the sample variance
    volatility = np.sqrt(sample_variance) * np.sqrt(260)  # Annualized volatility (assuming 260 trading days)
    
    return volatility



def get_rnr_values(df, period_dates_df, index_name):
    # Ensure index_name is not empty
    if not index_name:
        raise ValueError("Index name cannot be empty.")
    
    # Rename the 'Indexvalue' column dynamically using the provided index_name
    # df.rename(columns={'Indexvalue': f'Indexvalue_{index_name.upper()}'}, inplace=True)

    # Initialize an empty list to store results
    results = []

    # Loop over each period in period_dates_df
    for _, row in period_dates_df.iterrows():
        # Get the period and the corresponding start and end dates
        period = row['Period']
        start_date = row['Start Date']
        end_date = row['End Date']
        last_value_date = row['Last Value Date']
        
        # Filter the main dataframe to get the start value and end value for the return calculation
        start_data = df[df['Date'] == last_value_date]  # The value on the last value date
        end_data = df[df['Date'] == end_date]          # The value on the end date
        
        # Initialize return and volatility variables
        ann_return = None
        volatility = None
        
        # Calculate annualised return if valid data exists
        if not start_data.empty and not end_data.empty:
            start_value = start_data.iloc[0][f'Indexvalue_{index_name.upper()}']
            end_value = end_data.iloc[0][f'Indexvalue_{index_name.upper()}']
            ann_return = annualised_return(start_value, end_value, start_date, end_date, df)

        # Calculate the volatility for the period
        if not start_data.empty and not end_data.empty:
            volatility = calculate_volatility(last_value_date, end_date, df)

        # Store the results for this period with dynamic column names
        results.append({
            "Period": period,
            "Start Date": start_date,
            "End Date": end_date,
            f"Annualised Return_{index_name.upper()}": ann_return,
            f"Volatility_{index_name.upper()}": volatility
        })
    
    # Convert the results list into a DataFrame and return
    return pd.DataFrame(results)



def calculate_annual_returns(master_file, index_name, benchmark_name):
    # Ensure the 'Date' column is datetime
    master_file['Date'] = pd.to_datetime(master_file['Date'])
    
    # Extract the year from the 'Date' column
    master_file['Year'] = master_file['Date'].dt.year
    
    # Create empty lists to store the annual returns for the given index and benchmark
    annual_returns_index = []
    annual_returns_benchmark = []
    
    # Loop through each year in the dataset
    for year in master_file['Year'].unique():
        # Filter the data for the current year
        year_data = master_file[master_file['Year'] == year]
        
        # Get the starting value: last value from the previous year
        previous_year = year - 1
        previous_year_data = master_file[master_file['Year'] == previous_year]
        
        if not previous_year_data.empty:
            # Get the last available value from the previous year (Dec 31st or last available date)
            start_value_index = previous_year_data[f'Indexvalue_{index_name}'].iloc[-1]
            start_value_benchmark = previous_year_data[f'Indexvalue_{benchmark_name}'].iloc[-1]
            start_date = previous_year_data['Date'].iloc[-1]
        else:
            # If no data for the previous year, take the first value of the current year as start
            start_value_index = year_data[f'Indexvalue_{index_name}'].iloc[0]
            start_value_benchmark = year_data[f'Indexvalue_{benchmark_name}'].iloc[0]
            start_date = year_data['Date'].iloc[0]
        
        # Get the ending value: last available value of the current year
        end_value_index = year_data[f'Indexvalue_{index_name}'].iloc[-1]
        end_value_benchmark = year_data[f'Indexvalue_{benchmark_name}'].iloc[-1]
        end_date = year_data['Date'].iloc[-1]
        
        # Check if the year is complete
        december_data = year_data[year_data['Date'].dt.month == 12]
        is_complete = False
        
        if not december_data.empty:
            # If December data exists, check if January data for the next year exists
            january_data = master_file[(master_file['Year'] == year + 1) & (master_file['Date'].dt.month == 1)]
            if not january_data.empty:
                # If January data exists, the year is complete
                is_complete = True
        
        # Calculate the simple return for the current year
        return_index = (end_value_index - start_value_index) / start_value_index
        return_benchmark = (end_value_benchmark - start_value_benchmark) / start_value_benchmark
        
        # If the year is incomplete, annualize the return using available data
        if not is_complete:
            # Number of days available in the current year
            days_in_current_year = len(year_data)
            
            # Annualize the return using the formula: (1 + r)^(260 / days) - 1
            annualized_return_index = (1 + return_index) ** (260 / days_in_current_year) - 1
            annualized_return_benchmark = (1 + return_benchmark) ** (260 / days_in_current_year) - 1
        else:
            # For a full year, no annualization is required
            annualized_return_index = return_index
            annualized_return_benchmark = return_benchmark
        
        # Append the annual returns to the lists
        annual_returns_index.append((year, annualized_return_index))
        annual_returns_benchmark.append((year, annualized_return_benchmark))
    
    # Create a DataFrame with the calculated annual returns
    annual_returns_df = pd.DataFrame({
        'Year': [item[0] for item in annual_returns_index],
        f'Annual_return_{index_name}': [item[1] for item in annual_returns_index],
        f'Annual_return_{benchmark_name}': [item[1] for item in annual_returns_benchmark]
    })
    
    return annual_returns_df

def load_factor_data(file):
    # Load the Excel file using openpyxl
    wb = load_workbook(file, data_only=True)
    ws = wb['Factor Contributions']

    # Initialize an empty list to store the data
    data = []

    # Skip the first 6 rows and use the 7th row as column names
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 6:
            continue  # Skip the first 6 rows
        if all(cell is None for cell in row):  # Check if the entire row is empty
            break  # Stop once we hit an empty row

        # Append the row to data starting from the 7th row
        data.append(row)

    # Extract the 7th row (index 6) as column names
    column_names = data[0]  # The 7th row contains the column names

    # Create a DataFrame starting from the 8th row (index 1)
    factor_df = pd.DataFrame(data[1:], columns=column_names)
    factor_df['Source of Return'] = factor_df['Source of Return'].str.strip()  # Clean column names

    # Drop columns where the column name is None and all values are None
    factor_df = factor_df.dropna(axis=1, how='all')  # Drop columns where all values are NaN
    factor_df = factor_df.loc[:, factor_df.columns.notna()]  # Drop columns where column name is None

    # Return the relevant subset of the DataFrame
    return factor_df[['Source of Return', 'Contribution', 'Risk']]


def create_factor_lists(factor_df2):
    """
    This function takes a dataframe and generates lists based on the 'Source of Return' column.
    
    Args:
    factor_df2 (pd.DataFrame): The dataframe containing the factor data.
    
    Returns:
    dict: A dictionary containing lists for Style, Country, Industry, Currency, Market, and Sectors.
    """
    
    # Initialize lists to store the values
    style_list = []
    country_list = []
    industry_list = []
    currency_list = []
    market_list = []
    sector_list = []
    summary_list=['Portfolio','Benchmark','Active']
    decomp_list=['Style','Country','Industry','Currency','Market','Specific Return']

    # Start a flag to keep track of the section you're in
    current_section = None

    # Iterate over each row in the dataframe's 'Source of Return' column (or the column with text values)
    for value in factor_df2['Source of Return']:
        if value == "Style":
            current_section = "Style"
        elif value == "Country":
            current_section = "Country"
        elif value == "Industry":
            current_section = "Industry"
        elif value == "Currency":
            current_section = "Currency"
        elif value == "Market":
            current_section = "Market"
        elif value == "Sectors":
            current_section = "Sectors"
        
        # Add the value to the appropriate list based on the current section
        if current_section == "Style" and value != "Country":
            style_list.append(value)
            if value=='Style':
                style_list.remove(value)
        elif current_section == "Country" and value != "Industry":
            country_list.append(value)
            if value=='Country':
                country_list.remove(value)
        elif current_section == "Industry" and value != "Currency":
            industry_list.append(value)
            if value=='Industry':
                industry_list.remove(value)
        elif current_section == "Currency" and value != "Market":
            currency_list.append(value)
            if value=='Currency':
                currency_list.remove(value)
        elif current_section == "Market" and value != "Sectors":
            market_list.append(value)
            if value=='Market':
                market_list.remove(value)
        elif current_section == "Sectors":
            sector_list.append(value)
            if value=='Sectors':
                sector_list.remove(value)

    # Return the lists as a dictionary
    return {
        "Style": style_list,
        "Country": country_list,
        "Industry": industry_list,
        "Currency": currency_list,
        "Market": market_list,
        "Sectors": sector_list,
        "Summary":summary_list,
        "Return Decomposition":decomp_list
    }

# Function to load and display the factor attribution analysis
def factor_attribution_analysis(factor_file):
    # Load the factor data
    factor_data = load_factor_data(factor_file)  # Assuming a function to load the factor data
    
    # Create factor lists for Style, Country, Sector, Industry, Summary, and Decomposition
    factor_lists = create_factor_lists(factor_data)
    summary_list = factor_lists['Summary']
    decomp_list = factor_lists['Return Decomposition']
    style_list = factor_lists['Style']
    country_list = factor_lists['Country']
    sector_list = factor_lists['Sectors']
    industry_list = factor_lists['Industry']

    # Create DataFrames for Summary, Decomposition, Style, Country, Sector, and Industry
    summary_df = factor_data[factor_data['Source of Return'].isin(summary_list)]
    decomp_df = factor_data[factor_data['Source of Return'].isin(decomp_list)]
    decomp_df['Source of Return']=decomp_df['Source of Return'].replace('Specific Return','Stock Specific')
    style_df = factor_data[factor_data['Source of Return'].isin(style_list)]
    country_df = factor_data[factor_data['Source of Return'].isin(country_list)]
    sector_df = factor_data[factor_data['Source of Return'].isin(sector_list)]
    industry_df = factor_data[factor_data['Source of Return'].isin(industry_list)]


    ## Ordering for decomposition chart
    # # Manually define the new order of the 'Source of Return' values
    new_order = ['Style', 'Country', 'Industry', 'Currency', 'Market','Stock Specific']

    # Create a new column for the order
    decomp_df['New Order'] = pd.Categorical(decomp_df['Source of Return'], categories=new_order, ordered=True)

    # Sort the DataFrame based on the 'New Order' column
    decomp_df2 = decomp_df.sort_values(by='New Order').drop(columns=['New Order']) 

    ## Industry df reduced to TOP and BOTTOM 5
    industry_df2=industry_df.sort_values(by=['Contribution'],ascending=False)
    top5=industry_df2.head(5)
    bottom5=industry_df2.tail(5)
    industry_df3=pd.concat([top5,bottom5])
    # Function to create the vertical bar chart for summary
    def create_summary_vertical_fig(df, title):
        # Sort the DataFrame by 'Contribution' with positive contributions first
        df_sorted = df.sort_values(by=['Contribution'], ascending=False)
        # bar_colors = ['#000002' for val in df_sorted['Contribution']]  # All risk bars will have the same color

        # Create the vertical bar chart for Contribution
        bar_trace_contrib = go.Bar(
            x=df_sorted['Source of Return'],  # x for categories (Source of Return)
            y=df_sorted['Contribution'],      # y for the Contribution
            name='Contribution to Active Return',
            marker=dict(color='#000002'),  # Blue color for Contribution bars
            orientation='v'  # This makes the bars vertical
        )

        # Create the vertical bar chart for Risk
        bar_trace_risk = go.Bar(
            x=df_sorted['Source of Return'],  # x for categories (Source of Return)
            y=df_sorted['Risk'],      # y for the Risk
            name='Risk',
            marker=dict(color='#000001'),  # Orange color for Risk bars
            orientation='v'  # This makes the bars vertical
        )

        # Define the layout
        layout = go.Layout(
            title=title,
            yaxis=dict(
                tickformat='.1%',
                side='left', showgrid=True
            ),
            xaxis=dict(
                side='bottom', showgrid=False
            ),
            barmode='group',  # Group bars side by side
            showlegend=True,
            legend=dict(
                x=0.5,  # Position legend to the right of the plot
                y=-0.35,  # Position legend below the plot
                xanchor='center',  # Center the legend horizontally
                yanchor='bottom',
                traceorder='normal',
                orientation='h',
                font=dict(size=12),
                itemwidth=50,  # Set width for each legend item
                tracegroupgap=8
            )
        )

        # Create the figure
        fig = go.Figure(data=[bar_trace_contrib, bar_trace_risk], layout=layout)
        return fig
    # Function to create the horizontal bar chart for sector, industry, style, and country
    def create_fig(df, title):
        # Sort the DataFrame by 'Contribution' with positive contributions first
        df['Contribution_Sign'] = df['Contribution'] > 0  # Create a new column for sorting by sign
        df_sorted = df.sort_values(by=['Contribution_Sign', 'Contribution'], ascending=[False, False])
        bar_colors = ['#000002' if val > 0 else '#000001' for val in df_sorted['Contribution']]

        # Create the horizontal bar chart for Contribution
        bar_trace = go.Bar(
            y=df_sorted['Source of Return'],  # Change x to y for horizontal bars
            x=df_sorted['Contribution'],      # Change y to x for horizontal bars
            name='Contribution to Active Return',
            marker=dict(color=bar_colors),
            orientation='h'  # This makes the bars horizontal
        )

        # Create the scatter plot for Risk (with secondary x-axis)
        scatter_trace = go.Scatter(
            x=df_sorted['Risk'],  # Plotting 'Risk' on the secondary x-axis
            y=df_sorted['Source of Return'],  # Keeping 'Source of Return' on the y-axis
            mode='markers',
            name='Risk',
            xaxis='x2',  # Assign to secondary x-axis
            marker=dict(color='darkblue', size=6, line=dict(width=2, color='red'))
        )

        # Define the layout with two x-axes
        layout = go.Layout(
            title=title,
            xaxis=dict(
                title='Contribution to Active Return',
                tickformat='.3%',
                side='bottom', showgrid=False,ticks='outside'
            ),
            xaxis2=dict(  # Secondary x-axis for 'Risk'
                title='Risk',
                overlaying='x',
                tickformat='.3%',
                side='top',  # Put the secondary axis at the top
                showgrid=False,ticks='outside'
            ),
            yaxis=dict(
                side='left', showgrid=True
            ),
            barmode='group',
            showlegend=True
        )

        # Create the figure
        fig = go.Figure(data=[bar_trace, scatter_trace], layout=layout)
        return fig



   

    # Function to create the vertical bar chart for Decomposition
    def create_decomp_vertical_fig(df, title):

    # Create the vertical bar chart for Contribution
        bar_trace_contrib = go.Bar(
        x=df['Source of Return'],  # x for categories (Source of Return)
        y=df['Contribution'],      # y for the Contribution
        name='Contribution to Active Return',
        marker=dict(color='#000002'),
        orientation='v'  # This makes the bars vertical
    )

    # Create the vertical bar chart for Risk
        bar_trace_risk = go.Bar(
        x=df['Source of Return'],  # x for categories (Source of Return)
        y=df['Risk'],      # y for the Risk
        name='Risk',
        marker=dict(color='#000001'),  # Orange color for Risk bars
        orientation='v'  # This makes the bars vertical
    )

    # Define the layout
        layout = go.Layout(
        title=title,
        yaxis=dict(
            tickformat='.1%',
            side='left', showgrid=True
        ),
        xaxis=dict(
            side='bottom', showgrid=False
        ),
        barmode='group',  # Group bars side by side
        showlegend=True,
        legend=dict(
            x=0.5,  # Position legend to the right of the plot
            y=-0.35,  # Position legend below the plot
            xanchor='center',  # Center the legend horizontally
            yanchor='bottom',
            traceorder='normal',
            orientation='h',
            font=dict(size=12),
            itemwidth=50,  # Set width for each legend item
            tracegroupgap=8
        )
    )

    # Create the figure
        fig = go.Figure(data=[bar_trace_contrib, bar_trace_risk], layout=layout)
        return fig




    # Create the sector, industry, style, and country figures
    fig_sector = create_fig(sector_df, 'By Sector Factors')
    fig_industry = create_fig(industry_df3, 'By Industry Factors')
    fig_style = create_fig(style_df, 'By Style Factors')
    fig_country = create_fig(country_df, 'By Country Factors')

    # Create vertical bar charts for summary and decomp
    fig_summary = create_summary_vertical_fig(summary_df, 'Summary')
    fig_decomp = create_decomp_vertical_fig(decomp_df2, 'Contribution to Active Return: Factor Group Vs Stock Specific')

    return fig_sector, fig_industry, fig_style, fig_country, fig_summary, fig_decomp
